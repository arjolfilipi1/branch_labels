from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QPushButton, QTableWidget, 
                             QTableWidgetItem, QLabel, QMessageBox, QLineEdit,
                             QHeaderView,QStackedWidget,QComboBox,QSpinBox,QShortcut,
                             QDialog,QScrollArea,QFormLayout,QCheckBox,QDialogButtonBox,
                             QTextEdit, QFileDialog,QDoubleSpinBox)
from PyQt5.QtCore import Qt,QSettings,QLocale
from PyQt5 import QtGui
import os,sys
from PyQt5.QtPrintSupport import QPrinterInfo
import pyodbc
from PIL import Image,ImageWin
from reportlab.lib.pagesizes import mm
from reportlab.pdfgen import canvas
# from pylibdmtx.pylibdmtx import encode
from reportlab.lib.utils import ImageReader
from io import BytesIO
import tempfile,time
from PyQt5.QtGui import QClipboard,QPixmap,QKeySequence
import qrcode
import win32print,win32ui,win32api,win32con
import fitz  # PyMuPDF
import sqlite3
from operator import itemgetter
from openpyxl import load_workbook

# Application settings
SETTINGS = QSettings("Forschner", "Label EVOBUS")
locale = QLocale(QLocale.English)
QLocale.setDefault(locale)


class SettingsDialog(QDialog):
    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select File",
            "",  # Start in current directory
            "All Files (*);;Text Files (*.txt);;JSON Files (*.json)"
        )
        
        if file_path:
            self.file_path = file_path
            self.file_label.setText(os.path.basename(file_path))
            self.file_label.setToolTip(file_path)  # Show full path on hover
    def __init__(self, parent=None):
        super().__init__(parent)
        self.file_path = ""
        self.setWindowTitle("App Settings")
        self.resize(430, 600)
        # Create a scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        # Create a widget to hold the layout
        content_widget = QWidget()
        layout = QFormLayout(content_widget)
        
        # Database path
        self.file_label = QLabel("No file selected")
        self.browse_btn = QPushButton("Browse...")
        self.browse_btn.clicked.connect(self.browse_file)
        sqldb_layout = QHBoxLayout()
        sqldb_layout.addWidget(self.file_label)
        sqldb_layout.addWidget(self.browse_btn)
        
        # Database path
        self.db_path_edit = QLineEdit()
        self.db_path_edit.setPlaceholderText("Vendos Perdoruesin e xPPS")
        db_layout = QHBoxLayout()
        db_layout.addWidget(self.db_path_edit)
        
        # Images path
        self.password_edit = QLineEdit()
        self.password_edit.setPlaceholderText("Vendos passwordin e xPPS")
        password_layout = QHBoxLayout()
        password_layout.addWidget(self.password_edit)
        
        # dpi
        self.dpi_edit = QLineEdit()
        self.dpi_edit.setPlaceholderText("300")
        dpi_layout = QHBoxLayout()
        dpi_layout.addWidget(self.dpi_edit)
        
        # sql2
        self.sql = """
                SELECT BIMOD220.XBAU.XABNU2, BIMOD220.XBAU.XAMERK, BIMOD220.XBAU.XACDIL, BIMOD220.XBAU.XATENR,
    BIMOD220.XBAU.XACOMM, BIMOD220.XBAU.XASTAT, BIMOD220.XBAU.XAFIRM, BIMOD220.XBAU.XAWKNR FROM BIMOD220.XBAU
WHERE ( (BIMOD220.XBAU.XAMERK) IN ('3H','9T','DH','NT','VT')
        AND ((BIMOD220.XBAU.XABNU2) IN {} )
        AND (CAST(BIMOD220.XBAU.XASTAT AS INTEGER) < 90)
        AND ((BIMOD220.XBAU.XAFIRM) = '1')
        AND ((BIMOD220.XBAU.XAWKNR) = '000')
    )
            """
        self.con = "Driver={{IBM i Access ODBC Driver}};System=192.168.100.35;UID={};PWD={};DBQ=QGPL;"
        self.sql_edit = QTextEdit()
        self.sql_edit.setPlaceholderText(self.sql)
        slq_layout = QHBoxLayout()
        slq_layout.addWidget(self.sql_edit)
        
        self.con_edit = QTextEdit()
        self.con_edit.setPlaceholderText(self.con)
        con_layout = QHBoxLayout()
        con_layout.addWidget(self.con_edit)
        
        # sql2
        self.sql2 = """
    SELECT vodiče.VON AS Nga_dega, KABELY.Forsch_Nr_kabelu AS Moduli,vodiče.MODUL, vodiče.BIS AS Tek_dega,KABELY.výkresB0 AS B0,vodiče.Výroba AS Vyroba
FROM (KABELY 
INNER JOIN [KABELY NA POZICE] ON KABELY.Forsch_Nr_kabelu = [KABELY NA POZICE].Forsch_Nr_kabelu) 
INNER JOIN vodiče ON [KABELY NA POZICE].Pozice = vodiče.POS
WHERE KABELY.Forsch_Nr_kabelu IN ({});
    """
        self.con2 = "DSN=KomaxAL_Durres2;Driver={SQL Server};System=192.168.102.232;UID=komax;PWD=komax1;"
        
        self.sql2_edit = QTextEdit()
        self.sql2_edit.setPlaceholderText(self.sql2)
        slq2_layout = QHBoxLayout()
        slq2_layout.addWidget(self.sql2_edit)
        
        self.con2_edit = QTextEdit()
        self.con2_edit.setPlaceholderText(self.con2)
        con2_layout = QHBoxLayout()
        con2_layout.addWidget(self.con2_edit)
        
        
        # Add to form
        layout.addRow( sqldb_layout)
        layout.addRow("xPPS user:", db_layout)
        layout.addRow("Password:", password_layout)
        layout.addRow("dpi e printerit:", dpi_layout)
        layout.addRow("SQL e xpps:", slq_layout)
        layout.addRow("Conn e xpps:", con_layout)
        layout.addRow("SQL e fijeve:", slq2_layout)
        layout.addRow("Conn e fijeve:", con2_layout)
        
        # Dialog buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
        
        # Set the content widget to scroll area
        scroll.setWidget(content_widget)
        
        # Set scroll area as the main layout
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll)
        
        # Load current settings
        self.load_settings()

    def load_file(self):
        try:
            # Load from QSettings
            file_path = SETTINGS.value("file_path", "")
            
            # Update UI
            self.file_path = file_path
            if file_path:
                self.file_label.setText(os.path.basename(file_path))
                self.file_label.setToolTip(file_path)
            else:
                self.file_label.setText("No file selected")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load settings: {str(e)}")
    def load_settings(self):
        self.db_path_edit.setText(SETTINGS.value("xpps/user", "FILIPI"))
        self.password_edit.setText(SETTINGS.value("xpps/password", "a110033"))
        self.dpi_edit.setText(SETTINGS.value("app/dpi", "300"))
        self.sql_edit.setPlainText(SETTINGS.value("xpps/sql", self.sql))
        self.con_edit.setPlainText(SETTINGS.value("xpps/con", self.con))
        self.sql2_edit.setPlainText(SETTINGS.value("komax/sql", self.sql2))
        self.con2_edit.setPlainText(SETTINGS.value("komax/con", self.con2))
        self.load_file()
    def save_settings(self):
        SETTINGS.setValue("file_path", self.file_path)
        SETTINGS.setValue("xpps/user", self.db_path_edit.text())
        SETTINGS.setValue("xpps/password", self.password_edit.text())
        SETTINGS.setValue("app/dpi", self.dpi_edit.text())
        SETTINGS.setValue("xpps/sql", self.sql_edit.toPlainText())
        SETTINGS.setValue("xpps/con", self.con_edit.toPlainText())
        SETTINGS.setValue("komax/sql", self.sql2_edit.toPlainText())
        SETTINGS.setValue("komax/con", self.con2_edit.toPlainText())

class PrintSettingsDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Print Settings")
        self.resize(430, 600)
        # Create a scroll area
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        
        # Create a widget to hold the layout
        content_widget = QWidget()
        layout = QFormLayout(content_widget)
        
        self.font_combo = QComboBox()
        font_layout = QHBoxLayout()
        font_layout.addWidget(self.font_combo)
        
        self.printer_combo = QComboBox()
        pc_layout = QHBoxLayout()
        pc_layout.addWidget(self.printer_combo)
        
        # use qr
        self.checkbox = QCheckBox("Use qr", self)
        checkbox_layout = QHBoxLayout()
        checkbox_layout.addWidget(self.checkbox)
        
        # use qr
        self.s_print = QCheckBox("Use simple print", self)
        s_print_layout = QHBoxLayout()
        s_print_layout.addWidget(self.s_print)
        
        # dpi
        self.dpi_edit = QSpinBox()
        self.dpi_edit.setValue(300)
        self.dpi_edit.setMinimum(1)
        self.dpi_edit.setMaximum(1000000)
        dpi_layout = QHBoxLayout()
        dpi_layout.addWidget(self.dpi_edit)
        
        # min_font
        self.min_font_edit = QSpinBox()
        self.min_font_edit.setValue(6)
        self.min_font_edit.setMinimum(3)
        self.min_font_edit.setMaximum(25)
        min_font_layout = QHBoxLayout()
        min_font_layout.addWidget(self.min_font_edit)
        
        # max_font
        self.max_font_edit = QSpinBox()
        self.max_font_edit.setValue(25)
        self.max_font_edit.setMinimum(3)
        self.max_font_edit.setMaximum(25)
        max_font_layout = QHBoxLayout()
        max_font_layout.addWidget(self.max_font_edit)
        
        # width
        self.width_edit = QSpinBox()
        self.width_edit.setValue(50)
        self.width_edit.setMinimum(1)
        self.width_edit.setMaximum(1000)
        width_layout = QHBoxLayout()
        width_layout.addWidget(self.width_edit)
        
        # height
        self.height_edit = QSpinBox()
        # self.height_edit.setPlaceholderText("10")
        self.height_edit.setValue(10)
        self.height_edit.setMinimum(1)
        self.height_edit.setMaximum(1000)
        height_layout = QHBoxLayout()
        height_layout.addWidget(self.height_edit)
        
        # offset
        self.offset_edit = QSpinBox()
        # self.offset_edit.setPlaceholderText("2")
        self.offset_edit.setValue(2)
        self.offset_edit.setMinimum(-1)
        self.offset_edit.setMaximum(100)
        offset_layout = QHBoxLayout()
        offset_layout.addWidget(self.offset_edit)
        
        # horizontal offset
        self.h_off_edit = QDoubleSpinBox()
        self.h_off_edit.setValue(0.0)
        self.h_off_edit.setMinimum(-50)
        self.h_off_edit.setMaximum(50)
        self.h_off_edit.setSingleStep(0.1)
        h_offset_layout = QHBoxLayout()
        h_offset_layout.addWidget(self.h_off_edit)
        
        # vertikal offset
        self.v_off_edit = QDoubleSpinBox()
        self.v_off_edit.setValue(0.5)
        self.v_off_edit.setMinimum(-10)
        self.v_off_edit.setMaximum(10)
        self.v_off_edit.setSingleStep(0.1)
        v_offset_layout = QHBoxLayout()
        v_offset_layout.addWidget(self.v_off_edit)
        
        # Add to form
        layout.addRow("Font:", font_layout)
        layout.addRow("Zgjidh printerin:", pc_layout)
        layout.addRow("Perdor qr code:", checkbox_layout)
        layout.addRow("Perdor Printim automatik:", s_print_layout)
        layout.addRow("dpi e printerit:", dpi_layout)
        layout.addRow("gjeresi e etiketes (mm):", width_layout)
        layout.addRow("madhesia min e textit:", min_font_layout)
        layout.addRow("madhesia max e textit:", max_font_layout)
        layout.addRow("lartesia e etiketes (mm):", height_layout)
        layout.addRow("Fillimi i shkrimit (mm):", offset_layout)
        layout.addRow("Velizje horizontale (px):", offset_layout)
        
        # Dialog buttons
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        # layout.addRow(buttons)
        
        # Set the content widget to scroll area
        scroll.setWidget(content_widget)
        
        # Set scroll area as the main layout
        main_layout = QVBoxLayout(self)
        main_layout.addWidget(scroll)
        main_layout.addWidget(buttons)
        self.populate_printers()
        self.populate_fonts()
        # Load current settings
        self.load_settings()
    def populate_fonts(self):
        standard_fonts = [
    'Courier', 'Courier-Bold', 'Courier-Oblique', 'Courier-BoldOblique',
    'Helvetica', 'Helvetica-Bold', 'Helvetica-Oblique', 'Helvetica-BoldOblique',
    'Times-Roman', 'Times-Bold', 'Times-Italic', 'Times-BoldItalic',
    'Symbol', 'ZapfDingbats']       
        for font in standard_fonts:
            self.font_combo.addItem(font)
    def populate_printers(self):
        # Get all available printers
        printers = QPrinterInfo.availablePrinters()
        
        # Add printer names to combo box
        for printer in printers:
            self.printer_combo.addItem(printer.printerName())
            
        
            
    def load_settings(self):
        # Load saved printer from settings
        saved_printer = SETTINGS.value("saved_printer", "")
        try:
            if saved_printer:
            # Find the index of the saved printer
                index = self.printer_combo.findText(saved_printer)
                if index >= 0:
                    self.printer_combo.setCurrentIndex(index)
                    
        except:
            pass
        saved_font = SETTINGS.value("saved_font", "Courier-Bold")
        try:
            if saved_font:
            # Find the index of the saved printer
                index = self.font_combo.findText(saved_font)
                if index >= 0:
                    self.font_combo.setCurrentIndex(index)
                    
        except:
            pass
        if SETTINGS.value("simple_print", False) != 'false':
            self.s_print.setChecked(True)
        if SETTINGS.value("use_qr", False) != 'false':
            self.checkbox.setChecked(True)
        self.dpi_edit.setValue(int(SETTINGS.value("app/dpi", 300)))
        self.width_edit.setValue(int(SETTINGS.value("print/width", 49)))
        self.min_font_edit.setValue(int(SETTINGS.value("min_font", 6)))
        self.max_font_edit.setValue(int(SETTINGS.value("max_font", 25)))
        self.height_edit.setValue(int(SETTINGS.value("print/height", 10)))
        self.offset_edit.setValue(int(SETTINGS.value("print/offset", 2)))
    
    def save_settings(self):
        # Get selected printer
        selected_printer = self.printer_combo.currentText()
        selected_font = self.font_combo.currentText()
        index = self.printer_combo.findText(selected_printer)        
        # Save to settings
        SETTINGS.setValue("saved_printer", selected_printer)
        SETTINGS.setValue("saved_font", selected_font)
        SETTINGS.setValue("index_printer", index)
        SETTINGS.setValue("use_qr", self.checkbox.isChecked())
        SETTINGS.setValue("simple_print", self.s_print.isChecked())
        SETTINGS.setValue("app/dpi", str(self.dpi_edit.value()))
        SETTINGS.setValue("min_font", str(self.min_font_edit.value()))
        SETTINGS.setValue("max_font", str(self.max_font_edit.value()))
        SETTINGS.setValue("print/width", str(self.width_edit.value()))
        SETTINGS.setValue("print/height", str(self.height_edit.value()))
        SETTINGS.setValue("print/offset", str(self.offset_edit.value()))

class PartInventoryApp(QWidget):
    def __init__(self, db_name="inventory.db",parent = None):
        self.parent = parent
        super().__init__()
        # Database connection
        self.conn = None
        self.p1 = None
        self.db_name = file_path = SETTINGS.value("file_path", "")
        self.table_name = "Plan"
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        # Main widget and layout
        self.main_widget = QWidget()
        # self.setCentralWidget(self.main_widget)
        self.layout = QVBoxLayout(self)
        self.main_widget.setLayout(self.layout)
        
        # Database name input
        self.db_input_layout = QHBoxLayout()
        self.db_label = QLabel("Database Name:")
        self.db_input = QLineEdit(self.db_name)
        self.db_input_layout.addWidget(self.db_label)
        self.db_input_layout.addWidget(self.db_input)
        self.layout.addLayout(self.db_input_layout)
        
        
        
        # Buttons
        self.button_layout = QHBoxLayout()
        
        self.paste_button = QPushButton("Merr nga Excel")
        self.paste_button.clicked.connect(self.paste_from_clipboard)
        self.button_layout.addWidget(self.paste_button)
        
        self.add_row_button = QPushButton("Shto rresht")
        self.add_row_button.clicked.connect(self.add_empty_row)
        self.button_layout.addWidget(self.add_row_button)
        
        self.save_button = QPushButton("Update plan")
        self.save_button.clicked.connect(self.save_to_sqlite)
        self.button_layout.addWidget(self.save_button)
        
        self.clear_button = QPushButton("Delete plan")
        self.clear_button.clicked.connect(self.clear_table)
        self.button_layout.addWidget(self.clear_button)
        
        self.layout.addLayout(self.button_layout)
        
        # Table widget
        self.table = QTableWidget()
        self.table.setColumnCount(3)  # Part Number, Quantity, and Delete button columns
        self.table.setRowCount(0)
        self.table.setHorizontalHeaderLabels(["BB Number", "UG", "Action"])
        
        # Set column widths
        self.table.setColumnWidth(0, 300)  # Part Number
        self.table.setColumnWidth(1, 100)  # Quantity
        self.table.setColumnWidth(2, 80)   # Delete button
        
        # Allow editing for part number and quantity columns
        self.table.setEditTriggers(QTableWidget.AllEditTriggers)
        
        # Connect cell changed signal
        self.table.cellChanged.connect(self.on_cell_changed)
        
        # Track changes to prevent recursive signals
        self._processing_cell_change = False
        # Make the header resizeable
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        
        self.layout.addWidget(self.table)
        layout.addWidget(self.main_widget)
        # Status bar
        self.parent.statusBar().showMessage("Ready")
        
        # Initialize database
        self.init_db()
        self.refresh_table()
    def refresh_table(self):
        self.clear_table(add = False,delete = False)
        if not self.table_name:
            QMessageBox.warning(self, "Invalid Table Name", "Please enter a valid table name.")
            return
        # self.add_empty_row()  # Start with one empty row
        try:
            cursor = self.conn.cursor()
            cursor.execute(f"SELECT part_number,ug FROM {self.table_name};")
            for row in cursor.fetchall():
                self.table.cellChanged.disconnect(self.on_cell_changed)
                row_position = self.table.rowCount()
                self.table.insertRow(row_position)
                # Add empty items
                part_num_item = QTableWidgetItem(row[0])
                qty_item = QTableWidgetItem(row[1])
                self.table.setItem(row_position, 0, part_num_item)
                self.table.setItem(row_position, 1, qty_item)
                # Add delete button
                self.add_delete_button(row_position)
                self.table.cellChanged.connect(self.on_cell_changed)
        except Exception as e:
            QMessageBox.information(self, "setup error", f"Nuk eshte perzgjedhur folder me te dhenat:\n{str(e)}")
    def on_cell_changed(self, row, column):
        """Handle changes to table cells"""
        if self._processing_cell_change:
            return
            
        self._processing_cell_change = True
        
        try:
            item = self.table.item(row, column)
            if item:
                new_value = item.text()
                
                if column == 0:  # Part Number column
                    # Add your part number validation logic here
                    if '\n' in new_value:
                        self.delete_row(row)
                        self.paste_from_clipboard(new_value)
                elif column == 1:  # Quantity column
                    # Validate quantity is a number
                    pass
                
                # You could add auto-save functionality here if desired
                # self.save_to_sqlite()
                
                self.parent.statusBar().showMessage(f"Perditesim i rreshtit {row + 1}")
                
        finally:
            self._processing_cell_change = False
    def init_db(self):
        """Initialize or connect to SQLite database"""
        try:
            self.db_name = self.db_input.text() or "inventory.db"
            self.conn = sqlite3.connect(self.db_name)
            self.parent.statusBar().showMessage(f"Connected to database: {self.db_name}")
        except Exception as e:
            QMessageBox.critical(self, "Database Error", f"Nuk mundet te lidhet me databasen:\n{str(e)}")
    
    def paste_from_clipboard(self,text = None):
        """Paste data from clipboard into the table widget"""
        # Disconnect the cellChanged signal to prevent multiple triggers
        self.table.cellChanged.disconnect(self.on_cell_changed)
        clipboard = QApplication.clipboard()
        clipboard_text = clipboard.text()
        if text:
            clipboard_text = text
        if not clipboard_text.strip():
            QMessageBox.warning(self, "Empty Clipboard", "Ju lutem kopjoni te dhenat nga excel.")
            return
        
        try:
            # Split clipboard text into lines
            lines = clipboard_text.split('\n')
            
            # Filter out empty lines and clean whitespace
            part_numbers = [line.strip() for line in lines if line.strip()]
            
            if not part_numbers:
                QMessageBox.warning(self, "No Data", "Nuk jan kopjuar te dhena te sakta.")
                return
            
            # Get current row count
            current_rows = self.table.rowCount() if not text else self.table.rowCount() - 1 
            
            # Add new rows for each part number
            for i, part_num in enumerate(part_numbers):
                row_position = current_rows + i
                self.table.insertRow(row_position)
                
                # Set part number
                part_num_item = QTableWidgetItem(part_num)
                self.table.setItem(row_position, 0, part_num_item)
                
                qty_item = QTableWidgetItem("")
                self.table.setItem(row_position, 1, qty_item)
                
                # Add delete button
                self.add_delete_button(row_position)
            
            self.parent.statusBar().showMessage(f"U shtuan {len(part_numbers)} BB nr")
            
        except Exception as e:
            QMessageBox.critical(self, "Paste Error", f"Nuk mundet te lexohen te dhenat:\n{str(e)}")
        finally:
            self.table.cellChanged.connect(self.on_cell_changed)
    def add_empty_row(self):
        """Add an empty row at the end of the table"""
        self.table.cellChanged.disconnect(self.on_cell_changed)
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)
        
        # Add empty items
        part_num_item = QTableWidgetItem("")
        qty_item = QTableWidgetItem("")
        
        self.table.setItem(row_position, 0, part_num_item)
        self.table.setItem(row_position, 1, qty_item)
        
        # Add delete button
        self.add_delete_button(row_position)
        
        # Scroll to the new row
        self.table.scrollToBottom()
        self.table.cellChanged.connect(self.on_cell_changed)
    def add_delete_button(self, row):
        """Add a delete button to the specified row"""
        delete_button = QPushButton("Delete")
        delete_button.clicked.connect(lambda: self.delete_row(row))
        self.table.setCellWidget(row, 2, delete_button)
    
    def delete_row(self, row):
        """Delete the specified row from the table"""
        self.table.removeRow(row)
        
        # Update the delete buttons' connections for remaining rows
        for r in range(row, self.table.rowCount()):
            button = self.table.cellWidget(r, 2)
            if button:
                button.clicked.disconnect()
                button.clicked.connect(lambda _, r=r: self.delete_row(r))
        try:
            # Reconnect if needed
            if not self.conn:
                self.init_db()
            cursor = self.conn.cursor()
            part_item = self.table.item(row, 0)
            qty_item = self.table.item(row, 1)
            cursor.execute(f"DELETE FROM {self.table_name} WHERE part_number ='{part_item}' AND ug = '{qty_item}'")
            self.conn.commit()
            self.parent.statusBar().showMessage(f"Deleted row {row + 1}")
        except Exception as e:
                self.parent.statusBar().showMessage("XPPS delete plan", f"Could not save data to database:\n{str(e)}")
    
    def load_vod(self):
        try:
            # Reconnect if needed
            if not self.conn:
                self.init_db()
            
            try:
                cursor = self.conn.cursor()
                cursor.execute(f"SELECT XABNU2,XAMERK,XACDIL,XATENR AS Part_number,XACOMM FROM BB_query;")
                bb_data = cursor.fetchall()
                if not bb_data:
                    # self.parent.statusBar().showMessage("No BB_query data")
                    return None
                old_plan =[ "'" + str(r[3]) + "'" for r in bb_data ][:-1]
                old_plan =   ",".join(str(item) for item in old_plan) 
                vod_sql = (self.parent.SQL_TEMPLATE_VODICE.format(old_plan))
                vod_con = pyodbc.connect(self.parent.con2)
                vod_cursor = vod_con.cursor()
                vod_cursor.execute(vod_sql)
                v_data = {}
                cursor.execute("DELETE FROM ldata")
                for row in vod_cursor.fetchall():
                    part_num = row.Moduli.strip()
                    if part_num in v_data:
                        v_data[part_num].append(row)
                    else:
                        v_data[part_num]=[row]
                
                for r in bb_data:
                    if r[3].strip() in v_data:
                        v = v_data[r[3].strip()]
                        for vr in v:
                            cursor.execute(f"INSERT OR IGNORE INTO ldata (XABNU2,XAMERK,XACDIL,XATENR,XACOMM,VON,UniqueID,BB_UG,LabelID,výkresB0,Modul,Výroba) VALUES ('{r[0]}', '{r[1]}' ,'{r[2]}' ,'{r[3]}' ,'{r[4]}' ,'{vr[3]}' ,'{r[0].strip()}{r[1].strip()}{vr[3].strip()}' ,'{r[0].strip()}{r[1].strip()}','{vr[4].strip()}{r[1].strip()}{vr[3].strip()}','{vr[4]}','{vr[1]}','{vr[5]}')")
                            cursor.execute(f"INSERT OR IGNORE INTO ldata (XABNU2,XAMERK,XACDIL,XATENR,XACOMM,VON,UniqueID,BB_UG,LabelID,výkresB0,Modul,Výroba) VALUES ('{r[0]}', '{r[1]}' ,'{r[2]}' ,'{r[3]}' ,'{r[4]}' ,'{vr[0]}' ,'{r[0].strip()}{r[1].strip()}{vr[0].strip()}' ,'{r[0].strip()}{r[1].strip()}','{vr[4].strip()}{r[1].strip()}{vr[0].strip()}','{vr[4]}','{vr[1]}','{vr[5]}')")
                self.conn.commit()
                bb_list, bb_ug = self.p1.get_bb()
                self.p1.bb_ug = bb_ug
                self.p1.UG_combo.clear()
                self.p1.UG_combo.addItems(bb_ug[bb_list[0]])
                self.p1.BB_combo.clear()
                self.p1.BB_combo.addItems(bb_list)
                
                self.parent.statusBar().showMessage(f"Saved label data for {str(len(bb_data))} parts to table ")
            except Exception as e:
                self.conn.rollback()
                QMessageBox.critical(self, "Save Error1", f"second part:\n{str(e)}")
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Save Error", f"Could not save data to database:\n{str(e)}")
    
    def save_to_sqlite(self):
        """Save table data to SQLite database"""
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "No Data", "No data to save. Please add parts first.")
            return
        
        if not self.table_name:
            QMessageBox.warning(self, "Invalid Table Name", "Please enter a valid table name.")
            return
        
        try:
            # Reconnect if needed
            if not self.conn:
                self.init_db()
            
            cursor = self.conn.cursor()
            
            # Create table (drop if exists)
            # cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
            
            # Create new table with part_number and ug columns
            # cursor.execute(f"""
                # CREATE TABLE {table_name} (
                    # id INTEGER PRIMARY KEY AUTOINCREMENT,
                    # part_number TEXT,
                    # ug TEXT
                # )
# """)
            
            # Insert data (skip the delete button column)
            added_parts = 0
            bbs = []
            cursor.execute(f"SELECT part_number,ug FROM {self.table_name};")
            old_plan = [ r[0] for r in cursor.fetchall() ]
            
            for row in range(self.table.rowCount()):
                part_item = self.table.item(row, 0)
                qty_item = self.table.item(row, 1)
                
                part_number = part_item.text().strip() if part_item else ""
                quantity = qty_item.text().strip() if qty_item else ""
                bbs.append(part_number)
                
                if not part_number or part_number in old_plan:
                    continue  # Skip empty part numbers
                
                
                cursor.execute(
                    f"INSERT INTO {self.table_name} (part_number, ug) VALUES ('{part_number}', '{quantity}')"
                )
                added_parts += 1
            
            self.conn.commit()
            self.parent.statusBar().showMessage(f"Saved {added_parts} parts to table '{self.table_name}' in {self.db_name}")
            connection_string = self.parent.con.format(self.parent.user,self.parent.password)
            sql = self.parent.sql
            try:
                if bbs:
                    bbs_str = "("
                    for b in bbs:
                        bbs_str += "'"+b+"',"
                bbs_str = bbs_str[:-1] + ")"
                cursor.execute("DELETE FROM BB_query")
                xpps_conn = pyodbc.connect(connection_string)
                xpps_cursor = xpps_conn.cursor()
                sql = sql.format(bbs_str)
                xpps_cursor.execute(sql)
                for r in (xpps_cursor.fetchall()):
                    cursor.execute(f"INSERT OR IGNORE INTO BB_query (XABNU2,XAMERK,XACDIL,XATENR,XACOMM,XASTAT,XAFIRM,XAWKNR) VALUES ('{r[0]}', '{r[1]}' ,'{r[2]}' ,'{r[3]}' ,'{r[4]}' ,'{r[5]}' ,'{r[6]}' ,'{r[7]}')")
                    pass
                self.conn.commit()
                self.parent.load_proj_sep()
                self.parent.statusBar().showMessage("Successfully saved " + str(added_parts) + " parts to database")
                self.load_vod()
            except Exception as e:
                QMessageBox.critical(self, "XPPS BB_query", f"Could not save data to database BB_query:\n{str(e)}\n{connection_string}")
                
            
            
        except Exception as e:
            self.conn.rollback()
            QMessageBox.critical(self, "Save Error", f"Could not save data to database:\n{str(e)}")
    
    def clear_table(self,add=True,delete = True):
        """Clear the table widget"""
        self.table.clear()
        self.table.setRowCount(0)
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Part Number", "Quantity", "Action"])
        self.parent.statusBar().showMessage("Table cleared")
        if add:
            self.add_empty_row()  # Add one empty row back
        if delete:
            try:
                # Reconnect if needed
                if not self.conn:
                    self.init_db()
                
                cursor = self.conn.cursor()
                cursor.execute("DELETE FROM Plan")
                self.conn.commit()
                self.parent.statusBar().showMessage("Plani u fshi")
            except Exception as e:
                QMessageBox.critical(self, "Plan", f"Could not delete database Plan:\n{str(e)}")
    def closeEvent(self, event):
        """Clean up when closing the application"""
        if self.conn:
            self.conn.close()
        event.accept()

# ----------------------------
# Main Page (simple welcome screen)
# ----------------------------
class MainPage(QWidget):
    def __init__(self,conn,prt):
        self.ret = {}
        super().__init__()
        self.conn = conn
        self.bb = ""
        self.ug = ""
        self.bb_ug ={}
        self.BB_combo = None
        self.UG_combo = None
        self.prt = prt
        self.proj = []
        self.proj_sep = {}
        self.setup_ui()
        self.get_rend()
    def print_pdf_natively(self,pdf_path):
        if self.prt.simple_print:
            default_printer = win32print.GetDefaultPrinter()
            try:
                target_printer = SETTINGS.value("saved_printer", "")
                win32print.SetDefaultPrinter(target_printer)
                win32api.ShellExecute(0, "print", pdf_path, None, ".", 0)
            except Exception as e:
                self.prt.statusBar().showMessage(f"Sending to printing failed: {str(e)}")
            finally:
                win32print.SetDefaultPrinter(default_printer)
        else:
            doc = fitz.open(pdf_path)
            # Convert mm to pixels
            width_px = int((self.prt.label_width / 25.4) * self.prt.dpi)
            height_px = int((self.prt.label_height / 25.4) * self.prt.dpi)
            printer_name = SETTINGS.value("saved_printer", "")
            if not printer_name:
                printer_name = win32print.GetDefaultPrinter()
            page_width_px  = None
            page_height_px = None
            for i in range(len(doc)):
                try:
                    page = doc[i]
                    pix = page.get_pixmap(dpi=1600)
                    image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

                    self.prt.statusBar().showMessage(f"Printing page {i + 1}")

                    # Resize image to fixed size
                    resized = image.resize((width_px, height_px), Image.NEAREST)

                    # Create printer DC
                    hdc = win32ui.CreateDC()
                    hdc.CreatePrinterDC(printer_name)

                    # Start printing
                    hdc.StartDoc(f"PDF Page {i+1}")
                    hdc.StartPage()
                    if not page_width_px:
                        page_width_px = hdc.GetDeviceCaps(win32con.HORZRES)  # Printable area width
                        page_height_px = hdc.GetDeviceCaps(win32con.VERTRES)  # Printable area height
                    # Calculate center position
                    center_x = (page_width_px - width_px) // 2
                    center_y = (page_height_px - height_px) // 2
                    print(page_width_px,page_height_px,width_px,height_px)
                    print(center_x, center_y, center_x + width_px, center_y + height_px)
                    # Draw the image at 0,0 with 50x10mm size (in pixels)
                    dib = ImageWin.Dib(resized)
                    dib.draw(hdc.GetHandleOutput(), (center_x, center_y, center_x + width_px, center_y + height_px))

                    hdc.EndPage()
                    hdc.EndDoc()
                    hdc.DeleteDC()
                except Exception as e:
                    self.prt.statusBar().showMessage(f"Sending to printing failed: {str(e)} {i}")
    def print_pdf(self,data_list, output_pdf,nr_copies,use_qr):
        # Label dimensions
        label_width = self.prt.label_width  * mm
        label_height =self.prt.label_height * mm
        
        # Text area dimensions
        text_width = (self.prt.label_width - 10) * mm
        text_height = 9 * mm  # Available height for text
        
        # Data matrix dimensions
        dm_width = 10 * mm
        dm_height = 10 * mm
        # Create temporary PDF file (kept open to prevent deletion)
        temp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
        temp_pdf_path = temp_pdf.name
        temp_pdf.close()  # Close handle so Sumatra can access it
        try:
            # Create PDF
            c = canvas.Canvas(temp_pdf_path, pagesize=(label_width, label_height))
            temp_pdf = tempfile.NamedTemporaryFile(suffix='.pdf', delete=False)
            temp_pdf.close()
            for i in range(nr_copies):
            
                for index, item in enumerate(data_list):
                    # Start new page for each label
                    c.setPageSize((label_width, label_height))
                    
                    # TEXT (Left side)
                    font_size = self.prt.max_font
                    c.setFont(SETTINGS.value("saved_font", "Courier-Bold"), font_size)
                    
                    # Adjust font size if needed
                    while c.stringWidth(item, SETTINGS.value("saved_font", "Courier-Bold"), font_size) > ((text_width ) - (self.prt.label_offset*mm*2)) and font_size > self.prt.min_font:
                        font_size -= 0.5
                    
                    # Calculate text metrics for vertical centering
                    text_width_actual = c.stringWidth(item, SETTINGS.value("saved_font", "Courier-Bold"), font_size)
                    text_height_actual = font_size * 0.8  # Approximate text height
                    
                    # Calculate horizontal centering for text
                    text_x = self.prt.label_offset*mm
                    
                    # Calculate vertical centering - accounts for actual text height
                    vertical_offset = (label_height - text_height_actual) / 2
                    
                    # Draw centered text (both horizontally and vertically)
                    c.setFont(SETTINGS.value("saved_font", "Courier-Bold"), font_size)
                    c.drawString(text_x, vertical_offset, item)
                    if use_qr:
                        try:
                            qr = qrcode.QRCode(
                                version=1,
                                error_correction=qrcode.constants.ERROR_CORRECT_L,
                                box_size=10,  # Adjust this to control the size
                                border=4,
                            )
                            qr.add_data(item.encode('utf-8'))
                            qr.make(fit=True)

                            # Create PIL Image
                            img = qr.make_image(fill_color="black", back_color="white")
                            
                            # Prepare image for PDF
                            img_buffer = BytesIO()
                            img.save(img_buffer, format='PNG', dpi=(self.prt.dpi, self.prt.dpi))
                            img_buffer.seek(0)

                            # BARCODE (Right side) - Centered vertically with text
                            dm_x = label_width - dm_width - 1 * mm
                            dm_y = (label_height - dm_height) / 2  # Center vertically

                            # Add to PDF
                            c.drawImage(
                                ImageReader(img_buffer),
                                dm_x, dm_y,
                                width=dm_width,
                                height=dm_height,
                                preserveAspectRatio=True,
                                mask='auto'
                            )
                            
                            c.showPage()
                            
                        except Exception as e:
                            self.prt.statusBar().showMessage(f"Barcode generation failed: {str(e)}")
                    else:
                        try:
                                # Barcode
                            encoded = encode(item.encode('utf-8'))
                            img = Image.frombytes('RGB', (encoded.width, encoded.height), encoded.pixels)
                            img_buffer = BytesIO()
                            img.save(img_buffer, format='PNG')
                            img_buffer.seek(0)
                            
                            c.drawImage(ImageReader(img_buffer), 
                                       49*mm - 10*mm - 1*mm, 
                                       (9*mm - 8*mm)/2, 
                                       width=8*mm, 
                                       height=8*mm,
                                       preserveAspectRatio=True)
                            c.showPage()
                        except Exception as e:
                            self.prt.statusBar().showMessage(f"Barcode generation failed: {str(e)}")
            c.save()
            
            temp_pdf.close()
            time.sleep(0.5)
            try:
                self.print_pdf_natively(temp_pdf_path)
            # os.startfile(temp_pdf_path, 'print')
            except Exception as e:
                self.prt.statusBar().showMessage(f"Line 901 Printing failed before: {str(e)}")
        except Exception as e:
            self.prt.statusBar().showMessage(f"Line 903 Printing failed: {str(e)}")
        finally:
            try:
                os.unlink(temp_pdf.name)
            except:
                pass
    def setup_ui(self):
        # Main vertical layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # First row - image (centered in its own horizontal layout)
        image_row = QHBoxLayout()
        image_label = QLabel()
        pixmap = QPixmap("logo_forschner.png")  # Replace with your image path
        pixmap = pixmap.scaledToHeight(40, Qt.SmoothTransformation)
        image_label.setPixmap(pixmap)
        image_label.setAlignment(Qt.AlignCenter)
        image_row.addWidget(image_label)
        main_layout.addLayout(image_row)

        # Second row - "Skano BB" label (centered)
        title_row = QHBoxLayout()
        skano_label = QLabel("Skano BB")
        skano_label.setAlignment(Qt.AlignCenter)
        # skano_label.setStyleSheet("font-size: 20px; font-weight: bold;")
        title_row.addWidget(skano_label)
        main_layout.addLayout(title_row)

        # Third row - info label (centered)
        info_row = QHBoxLayout()
        info_label = QLabel("BB Number:")
        info_label.setAlignment(Qt.AlignCenter)
        info_row.addWidget(info_label)
        main_layout.addLayout(info_row)

        # Fourth row - text input with stretch on both sides
        input_row = QHBoxLayout()
        input_row.addStretch()  # Add stretch before
        self.bb_input = QLineEdit()
        self.bb_input.setPlaceholderText("Skano BB")
        self.bb_input.setFixedWidth(300)  # Set a reasonable width
        self.bb_input.returnPressed.connect(self.print_bb)
        input_row.addWidget(self.bb_input)
        input_row.addStretch()  # Add stretch after
        main_layout.addLayout(input_row)

        # Fifth row - button (centered)
        button_row = QHBoxLayout()
        button_row.addStretch()
        sub_button = QPushButton("Submit")
        sub_button.setStyleSheet("""
            QPushButton {
                padding: 8px;
                font-size: 16px;
                min-width: 100px;
            }
        """)
        button_row.addWidget(sub_button)
        sub_button.clicked.connect(self.print_bb)
        button_row.addStretch()
        main_layout.addLayout(button_row)
        
        # First row - Single label
        first_row_label = QLabel("Printo etiketat ne baze te BB dhe UG")
        first_row_label.setAlignment(Qt.AlignCenter)
        first_row_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                padding: 10px;
                background-color: #f0f0f0;
                border-radius: 5px;
            }
        """)
        
        # Second row - Three columns
        second_row_layout = QHBoxLayout()
        second_row_layout.setSpacing(20)  # Space between columns
        if ( self.get_bb()):
            bb_list,self.bb_ug = self.get_bb()
            if bb_list == [] or self.bb_ug == None: 
                pass
            else:
                self.bb = bb_list[0]
                self.ug = self.bb_ug[bb_list[0]][0]
        else:
            bb_list = []
            self.bb_ug = {}
            self.ug = []
        # Create three columns
        columns_data = [
            {"label": "Zgjidh BB", "items": [b for b in bb_list] if bb_list else [""] },
            {"label": "Zgjidh UG", "items": self.bb_ug[bb_list[0]] if bb_list else [""]},

        ]
        
        self.combo_boxes = []  # Store combo boxes for later access
        
        for i, column_data in enumerate(columns_data):
            # Create a vertical layout for each column
            column_layout = QVBoxLayout()
            column_layout.setSpacing(8)
            
            # Create label for the column
            label = QLabel(column_data["label"])
            label.setAlignment(Qt.AlignCenter)
            label.setStyleSheet("""
                QLabel {
                    font-weight: bold;
                    color: #333333;
                }
            """)
            
            # Create combo box (drop down)
            combo_box = QComboBox()
            if i==0:
                self.BB_combo = combo_box 
                combo_box.setObjectName("BB_combo" )
            else:
                self.UG_combo = combo_box
                combo_box.setObjectName("UG_combo" )
            combo_box.addItems(column_data["items"])
            combo_box.setMinimumWidth(120)  # Set minimum width for consistency
            
            
            
            # Connect signal for selection changes
            combo_box.currentIndexChanged.connect(
                lambda index, cb=combo_box: self.on_combo_changed(cb, index)
            )
            
            # Add widgets to column layout
            column_layout.addWidget(label)
            column_layout.addWidget(combo_box)
            
            # Add column layout to second row
            second_row_layout.addLayout(column_layout)
            
            self.combo_boxes.append(combo_box)
        column_layout = QVBoxLayout()
        column_layout.setSpacing(8)
        BBUG_button = QPushButton("Printo etiketa")
        BBUG_button.setStyleSheet("""
            QPushButton {
                padding: 8px;
                font-size: 16px;
                min-width: 100px;
            }
        """)
        column_layout.addWidget(BBUG_button)
        second_row_layout.addLayout(column_layout)
        
        BBUG_button.clicked.connect(self.print_bbug)
        # Add rows to main layout
        bordered_widget = QWidget()
        bordered_widget.setObjectName("bo")
        
        # Apply border styling to the widget
        bordered_widget.setStyleSheet("""
            #bo {
                border: 2px solid #cccccc;
                border-radius: 5px;
                padding: 5px;
                background-color: #ffffff;
            }
        """)
        third_row_layout = QHBoxLayout(bordered_widget)
        third_row_layout.setSpacing(20)  # Space between columns
        # Column layout
        column_layout = QVBoxLayout()
        column_layout.setSpacing(15)
        
        # "Separate" label
        separate_label = QLabel("Separate")
        separate_label.setAlignment(Qt.AlignCenter)
        column_layout.addWidget(separate_label)
        
        # "Projekti" label
        projekti_label = QLabel("Projekti")
        projekti_label.setAlignment(Qt.AlignCenter)
        column_layout.addWidget(projekti_label)
        
        # Dropdown for Projekti
        self.projekti_combo = QComboBox()
        column_layout.addWidget(self.projekti_combo)
        self.projekti_combo.currentIndexChanged.connect(
                lambda index, cb=self.projekti_combo: self.on_projenti_changed(cb, index)
            )
        # "separati" label
        separati_label = QLabel("separati")
        separati_label.setAlignment(Qt.AlignCenter)
        column_layout.addWidget(separati_label)
        
        # Combo box for separati
        self.separati_combo = QComboBox()
        self.separati_combo.setEditable(True)
        column_layout.addWidget(self.separati_combo)
        self.load_proj_sep()
        # "sasi" label
        separati_label = QLabel("nr i kopjeve")
        separati_label.setAlignment(Qt.AlignCenter)
        column_layout.addWidget(separati_label)
        
        self.sasi_edit = QSpinBox()
        self.sasi_edit.setLocale(locale)
        self.sasi_edit.setValue(1)
        self.sasi_edit.setMinimum(1)
        self.sasi_edit.setMaximum(1000000)
        column_layout.addWidget(self.sasi_edit)
        # Button
        process_button = QPushButton("Printo etiketa")
        
        process_button.clicked.connect(self.separate_clicked)
        column_layout.addWidget(process_button)
        third_row_layout.addLayout(column_layout,1)
        
        column_layout = QVBoxLayout()
        column_layout.setSpacing(15)
        
        # "Etiketa shtese" label
        etiketa_label = QLabel("Etiketa shtese")
        etiketa_label.setAlignment(Qt.AlignCenter)
        column_layout.addWidget(etiketa_label)
        
        # Text input
        self.text_input = QLineEdit()
        self.text_input.setPlaceholderText("Vendos degen")
        column_layout.addWidget(self.text_input)
        
        # Button
        add_button = QPushButton("Printo nje etikete")
        add_button.clicked.connect(self.teke)
        column_layout.addWidget(add_button)
        
        third_row_layout.addLayout(column_layout,1)
        # main_layout.addWidget(first_row_label)
        # main_layout.addLayout(second_row_layout)
        
        main_layout.addWidget(bordered_widget)
    def load_proj_sep(self):
        self.separati_combo.clear()
        self.projekti_combo.clear()
        self.proj, self.proj_sep = self.get_proj() if self.get_proj() else [[],{}]
        if self.proj:
            self.projekti_combo.addItems(self.proj)
            self.separati_combo.addItems(self.proj_sep[self.proj[0]])
    def separate_clicked(self):
        project = self.projekti_combo.currentText().strip()
        option = self.separati_combo.currentText().strip()
        quant = self.sasi_edit.value()
        try:
            dl =(self.ret[project][option][::-1])
            self.print_pdf( sorted(dl) + [option.strip()] ,'text.pdf',quant,self.prt.use_qr)
        except Exception as e:
            QMessageBox.critical(self, "Separati nuk eshte regjistruar", f"Nuk eshte regjistruar separati per kete projekt:\n{str(e)}")
        return None
        
    def teke(self):
        text = self.text_input.text()
        if text:
            self.text_input.clear()
            self.print_pdf([text],'text.pdf',1,self.prt.use_qr)

    def print_bb(self):
        bb_sl_ug = self.bb_input.text().strip()
        if "/" not in bb_sl_ug:
            QMessageBox.critical(self, "BB print", "Barkodi i skanuar nuk i pershtatet formatit 'bb/ug'")
            return None
        bb,ug = bb,ug = bb_sl_ug.split("/")
        try:
            connection_string = self.prt.con.format(self.prt.user,self.prt.password)
            xpps_conn = pyodbc.connect(connection_string)
            xpps_cursor = xpps_conn.cursor()
            sql = self.prt.sql.format("'" + bb + "'")
            xpps_cursor.execute(sql)
            pn_list = []
            for r in (xpps_cursor.fetchall()):
                if r[1].strip() == ug.strip():
                    pn_list.append(r[3])
            if not pn_list:
                QMessageBox.critical(self, "UG error", f"Lista per kete bb/ug eshte bosh")
                return
            old_plan =[ "'" + str(r) + "'" for r in pn_list ]
            old_plan =   ",".join(str(item) for item in old_plan)
            vod_sql = (self.prt.SQL_TEMPLATE_VODICE.format(old_plan))
            vod_con = pyodbc.connect(self.prt.con2)
            vod_cursor = vod_con.cursor()
            vod_cursor.execute(vod_sql)
            ts = []
            data_list = []
            for row in vod_cursor.fetchall():
                if row[0] not in data_list:
                    data_list.append(row[0])
                    rid = row[4].strip() + row[5].strip() + row[0].strip()
                    if rid in self.rend:
                        r = self.rend[rid]
                        ts.append( [r[0],r[1],r[2], row[0]] )
                    else:
                        ts.append( ["zzzzz",9999,9999, row[0]] )
                if row[3] not in data_list:
                    data_list.append(row[3])
                    rid = row[4].strip() + row[5].strip() + row[3].strip()
                    if rid in self.rend:
                        r = self.rend[rid]
                        ts.append( [r[0],r[1],r[2], row[3]] )
                    else:
                        ts.append( ["zzzzz"+row[3],9999,9999, row[3]] )
            if data_list:
                ts = sorted(ts, key=itemgetter(0, 1, 2))
                print([t[3] for t in ts])
                self.print_pdf( ([t[3] for t in ts]) + [bb.strip()],'text.pdf',1,self.prt.use_qr)
        except Exception as e:
                QMessageBox.critical(self, "Xpps connection", f"Nuk mund te lidhemi me XPPS:\n{str(e)}")
        
    
    def print_bbug(self):
        qry = "SELECT ldata.VON FROM ldata WHERE UniqueID LIKE ('"+self.bb.strip()+self.ug.strip()+"%');"
        if self.conn:
            try:
                data_list = []
                cursor = self.conn.cursor()
                cursor.execute(qry)
                von_data = cursor.fetchall()
                if not von_data:
                    self.prt.statusBar().showMessage("No BB_query data")
                    return None
                for v in von_data:
                    data_list.append(v[0])
                self.print_pdf( sorted(data_list) + [self.bb.strip(),self.ug.strip()],'text.pdf',1,self.prt.use_qr)
                #add print statement here
            except Exception as e:
                QMessageBox.critical(self, "BB print", f"Could not connect to database for bb ug search:\n{str(e)}")
    def on_projenti_changed(self, combo_box, index):
        
        selected_text = combo_box.currentText()
        self.separati_combo.clear()
        self.separati_combo.addItems(self.proj_sep[selected_text])
    def on_combo_changed(self, combo_box, index):
        """Handle combo box selection changes"""
        selected_text = combo_box.currentText()
        if combo_box.objectName() == "BB_combo" and selected_text:
            self.bb = selected_text
            self.UG_combo.clear()
            self.UG_combo.addItems(self.bb_ug[selected_text])
            self.ug = self.bb_ug[selected_text][0]
        elif selected_text:
            self.ug = selected_text
    
    def get_rend(self):
        file_path = SETTINGS.value("file_path", "")
        if file_path:
            file_path = file_path.replace("inventory.db","CitaroLabelsort.xlsx")
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet = workbook.active
            rend = {}
            i = 0
            for row in sheet:
                i +=1
                if i ==1 or row[0].value == None:
                    continue
                else:
                    try:
                        rid = row[0].value.strip() 
                        r1 =  row[4].value.strip()
                        r2 =  row[5].value.strip()
                        ar =  row[6].value.strip()
                        rend[rid] = [ar,r1,r2]
                    except:
                        pass
            self.rend = rend
    def get_proj(self):
        file_path = SETTINGS.value("file_path", "")
        if file_path:
            file_path = file_path.replace("inventory.db","Etiketat_Separate.xlsx")
            workbook = load_workbook(filename=file_path, read_only=True)
            sheet = workbook.active
            ret = {}
            data_list = []
            bb_ug = {}
            i = 0
            for row in sheet:
                i += 1
                if i == 1 or row[5].value == None:
                    continue
                p = row[5].value.strip() if row[5].value else None
                d = row[2].value.strip() if row[2].value else None
                n = row[1].value.strip() if row[1].value else None
                if p not in ret:
                    ret[p] = {}
                    ret[p][n] = [d]
                    data_list.append(p)
                elif n  not in ret[p]:
                    ret[p][n] = [d]
                if p in ret and n in ret[p] and d not in ret[p][n]:
                    ret[p][n].append(d)
                if p not in bb_ug:
                    bb_ug[p] = [n]
                elif n not in bb_ug[p]:
                    bb_ug[p].append(n)
            self.ret = ret
            return data_list,bb_ug
        
    def get_bb(self):
        if self.conn:
            data_list = []
            bb_ug = {}
            try:
                cursor = self.conn.cursor()
                cursor.execute("SELECT DISTINCT ldata.XABNU2, ldata.XAMERK FROM ldata;")
                von_data = cursor.fetchall()
                if not von_data:
                    self.prt.statusBar().showMessage("No BB_query data")
                    return None
                for v in von_data:
                    if v[0] not in bb_ug:
                        bb_ug[v[0]] = [v[1]]
                    elif v[1] not in bb_ug[v[0]]:
                        bb_ug[v[0]].append(v[1])
                    if v[0] not in data_list:
                        data_list.append(v[0])
                return data_list,bb_ug
            except Exception as e:
                QMessageBox.critical(self, "get BB", f"Could not connect to database :\n{str(e)}")
                return data_list,bb_ug
    
# ----------------------------
# Main Application Window
# ----------------------------
class MainApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Barcode printing Evo Bus")
        self.setGeometry(50, 50, 400, 400)
        self.sql = None
        self.conn = None
        self.SQL_TEMPLATE_VODICE = None
        self.con2 = None
        self.min_font = int(SETTINGS.value("min_font", 6))
        self.max_font = int(SETTINGS.value("max_font", 25))
        self.index_printer = int(SETTINGS.value("index_printer", 0))
        self.dpi = int(SETTINGS.value("app/dpi", 300))
        self.label_width = int(SETTINGS.value("print/width", 50))
        self.label_height = int(SETTINGS.value("print/height", 10))
        self.label_offset = int(SETTINGS.value("print/offset", 2))
        self.dpi = int(SETTINGS.value("app/dpi", 300))
        self.use_qr = SETTINGS.value("use_qr", False) == 'true'
        self.simple_print = SETTINGS.value("simple_print", False) == 'true'
        self.user = (SETTINGS.value("xpps/user", "FILIPI"))
        self.password = (SETTINGS.value("xpps/password", "a110033"))
        print_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        print_shortcut.activated.connect(self.p_setlings)
        con_shortcut = QShortcut(QKeySequence("Ctrl+Shift+D"), self)
        con_shortcut.activated.connect(self.d_setlings)
        self.get_settings()
        self.setup_ui()
        self.appl_style()
    def get_settings(self):
        self.sql = SETTINGS.value("xpps/sql", """"
                 SELECT BIMOD220.XBAU.XABNU2, BIMOD220.XBAU.XAMERK, BIMOD220.XBAU.XACDIL, BIMOD220.XBAU.XATENR,
    BIMOD220.XBAU.XACOMM, BIMOD220.XBAU.XASTAT, BIMOD220.XBAU.XAFIRM, BIMOD220.XBAU.XAWKNR FROM BIMOD220.XBAU
WHERE ( (BIMOD220.XBAU.XAMERK) IN ('3H','9T','DH','NT','VT')
        AND ((BIMOD220.XBAU.XABNU2) IN {} )
        AND (CAST(BIMOD220.XBAU.XASTAT AS INTEGER) < 90)
        AND ((BIMOD220.XBAU.XAFIRM) = '1')
        AND ((BIMOD220.XBAU.XAWKNR) = '000')
    )""")
        self.con = SETTINGS.value("xpps/con", "Driver={{IBM i Access ODBC Driver}};System=192.168.100.35;UID={};PWD={};DBQ=QGPL;")
        self.SQL_TEMPLATE_VODICE = SETTINGS.value("komax/sql", """
SELECT vodiče.VON AS Nga_dega, KABELY.Forsch_Nr_kabelu AS Moduli,vodiče.MODUL, vodiče.BIS AS Tek_dega,KABELY.výkresB0 AS B0,KABELY.Výroba AS Vyroba
FROM (KABELY 
INNER JOIN [KABELY NA POZICE] ON KABELY.Forsch_Nr_kabelu = [KABELY NA POZICE].Forsch_Nr_kabelu) 
INNER JOIN vodiče ON [KABELY NA POZICE].Pozice = vodiče.POS
WHERE KABELY.Forsch_Nr_kabelu IN ({});
""")
        self.con2 = SETTINGS.value("komax/con", "DSN=KomaxAL_Durres2;Driver={SQL Server};System=192.168.102.232;UID=komax;PWD=komax1;")
        self.use_qr = SETTINGS.value("use_qr", False) == 'true'
        self.dpi = int(SETTINGS.value("app/dpi", 300))
        self.min_font = int(SETTINGS.value("min_font", 6))
        self.max_font = int(SETTINGS.value("max_font", 25))
        self.index_printer = int(SETTINGS.value("index_printer", 0))
        self.label_width = int(SETTINGS.value("print/width", 50))
        self.label_height = int(SETTINGS.value("print/height", 10))
        self.label_offset = int(SETTINGS.value("print/offset", 2))
        self.user = (SETTINGS.value("xpps/user", "FILIPI"))
        self.password = (SETTINGS.value("xpps/password", "a110033"))
    def appl_style(self):
        self.setStyleSheet("QHBoxLayout {border: 2px solid #d40000; border-radius: 5px;} QVBoxLayout {border: 2px solid #d40000; border-radius: 5px;} QSpinBox { background-color: white; color: #d40000; border: 2px solid #d40000; border-radius: 5px; padding: 5px; font-size: 12px; min-width: 100px; } QWidget { background-color: white; font-family: Arial; font-size: 14px; } QLabel { color: #b00000; font-weight: bold; text-align: center;  } QPushButton { background-color: #b00000; color: white; border: none; padding: 3px 6px; border-radius: 4px; } QPushButton:hover { background-color: #d00000; } QPushButton:pressed { background-color: #900000; } QLineEdit, QComboBox { border: 1px solid #b00000; border-radius: 4px; padding: 4px; background-color: #fff0f0; text-align: center;} QListWidget { border: 1px solid #b00000; border-radius: 6px; background-color: #fff4f4; padding: 4px; } QListWidget::item { padding: 6px; border-bottom: 1px solid #ffd6d6; } QListWidget::item:selected { background-color: #ffcccc; color: #900000; font-weight: bold; border: 1px solid #b00000; } QScrollBar:vertical { border: none; background: #fff0f0; width: 10px; margin: 2px 0 2px 0; } QScrollBar::handle:vertical { background: #d00000; min-height: 20px; border-radius: 5px; } QScrollBar::handle:vertical:hover { background: #b00000; } QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }")
    def p_setlings(self):
        dialog = PrintSettingsDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            dialog.save_settings()
            self.get_settings()
    def d_setlings(self):
        dialog = SettingsDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            dialog.save_settings()
            self.get_settings()
    def setup_ui(self):
        self.setWindowIcon(QtGui.QIcon('logo.ico'))
        self.SQL_TEMPLATE_VODICE = SETTINGS.value("komax/sql","""
SELECT vodiče.VON AS Nga_dega, KABELY.Forsch_Nr_kabelu AS Moduli,vodiče.MODUL, vodiče.BIS AS Tek_dega,KABELY.výkresB0 AS B0,vodiče.Výroba AS Vyroba
FROM (KABELY 
INNER JOIN [KABELY NA POZICE] ON KABELY.Forsch_Nr_kabelu = [KABELY NA POZICE].Forsch_Nr_kabelu) 
INNER JOIN vodiče ON [KABELY NA POZICE].Pozice = vodiče.POS
WHERE KABELY.Forsch_Nr_kabelu IN ({});
""")
        
        # Central Widget and Layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Sidebar Navigation
        sidebar = QWidget()
        sidebar.setFixedWidth(150)
        sidebar_layout = QVBoxLayout(sidebar)
        
        self.btn_main = QPushButton("Main Page")
        self.btn_inventory = QPushButton("Plan")
        
        sidebar_layout.addWidget(self.btn_main)
        sidebar_layout.addWidget(self.btn_inventory)
        sidebar_layout.addStretch()

        # Page Container (Stacked Widget)
        self.pages = QStackedWidget()
        self.p_two = PartInventoryApp("inventory.db",parent = self)
        p1 = MainPage(self.p_two.conn,self)
        self.load_proj_sep = p1.load_proj_sep
        self.pages.addWidget(p1) # Index 0
        self.p_two.p1 = p1
        self.pages.addWidget(self.p_two)  # Index 1

        # Add to main layout
        main_layout.addWidget(sidebar)
        main_layout.addWidget(self.pages)

        # Connect navigation buttons
        self.btn_main.clicked.connect(lambda: self.pages.setCurrentIndex(0))
        self.btn_inventory.clicked.connect(self.page_two)
        
    def page_two(self):
        self.pages.setCurrentIndex(1)
        self.p_two.refresh_table()
# ----------------------------
# Run the Application
# ----------------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainApp()
    window.show()
    sys.exit(app.exec_())